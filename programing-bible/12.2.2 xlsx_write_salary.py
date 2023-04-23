# python 建立 excel 薪資表

import openpyxl

# 建立新的 Excel 檔案
workbook = openpyxl.Workbook()

# 選取預設的工作表
sheet = workbook.active

# 設定表頭資料
sheet['A1'] = '員工編號'
sheet['B1'] = '姓名'
sheet['C1'] = '基本薪資'
sheet['D1'] = '津貼'
sheet['E1'] = '總薪資'

# 設定員工資料
employee_data = [
    {'員工編號': '001', '姓名': '小明', '基本薪資': 30000, '津貼': 5000, '總薪資': 35000},
    {'員工編號': '002', '姓名': '小華', '基本薪資': 28000, '津貼': 4000, '總薪資': 32000},
    {'員工編號': '003', '姓名': '小英', '基本薪資': 32000, '津貼': 5500, '總薪資': 37500},
]

# 寫入員工資料
for i, data in enumerate(employee_data, start=2):
    sheet.cell(row=i, column=1).value = data['員工編號']
    sheet.cell(row=i, column=2).value = data['姓名']
    sheet.cell(row=i, column=3).value = data['基本薪資']
    sheet.cell(row=i, column=4).value = data['津貼']
    sheet.cell(row=i, column=5).value = data['總薪資']

# 設定列寬
sheet.column_dimensions['A'].width = 12
sheet.column_dimensions['B'].width = 12
sheet.column_dimensions['C'].width = 12
sheet.column_dimensions['D'].width = 12
sheet.column_dimensions['E'].width = 12

# 儲存檔案
workbook.save('salary.xlsx')